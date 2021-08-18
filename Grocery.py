import openpyxl
wb=openpyxl.load_workbook("C:\\Users\\Intel\\Desktop\\grocerry.xlsx")
sheets=wb.sheetnames
sh1=wb['Sheet1']
print('The available items are')
print('FRUITS\nVEGETABLES\nSNACKS')
num=int(input('Enter the number of items you want'))
for i in range(0,num):
    choice=input('Enter your choices from\n\t\tFRUITS\t:\tVEGETABLES\t:\tSNACKS\n')
    if(choice=='fruits'):
        print('\tAvailable fruits\t')
        for i in range(2,12):
            b=sh1.cell(i,1).value
            print(b.upper())
        n=int(input('Enter the number of fruits\t:\t'))
        for i in range(0,n):
            print('Enter fruit name', i+1)
            m=input()
            f=m.lower()
            for i in range(2,12):
                d=sh1.cell(i,1).value
                if f in d:
                    w=sh1.cell(i,2).value
                    print('weight\t:\t',w)
                    c=sh1.cell(i,3).value
                    print('cost\t:\t',c,'Rupees')
                    l=sh1.cell(i,4).value
                    print('Location:\t',l)
    if(choice=='vegetables'):
        print('\tAvailable vegetables\t')
        for i in range(14,24):
            b=sh1.cell(i,1).value
            print(b.upper())
        n=int(input('Enter the number of vegetables\t:\t'))
        for i in range(0,n):
            print('Enter vegetable name', i+1)
            m=input()
            f=m.lower()
            for i in range(14,24):
                d=sh1.cell(i,1).value
                if f in d:
                    w=sh1.cell(i,2).value
                    print('weight\t:\t',w)
                    c=sh1.cell(i,3).value
                    print('cost\t:\t',c,'Rupees')
                    l=sh1.cell(i,4).value
                    print('Location:\t',l)
    if(choice=='snacks'):
        print('\tAvailable snacks\t')
        for i in range(26,36):
            b=sh1.cell(i,1).value
            print(b.upper())
        n=int(input('Enter the number of snack\t:\t'))
        for i in range(0,n):
            print('Enter snack name', i+1)
            m=input()
            f=m.lower()
            for i in range(26,36):
                d=sh1.cell(i,1).value
                if f in d:
                    w=sh1.cell(i,2).value
                    print('weight\t:\t',w)
                    c=sh1.cell(i,3).value
                    print('cost\t:\t',c,'Rupees')
                    l=sh1.cell(i,4).value
                    print('Location:\t',l)

    
                    
